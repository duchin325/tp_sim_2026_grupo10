import openpyxl
import sys
sys.stdout.reconfigure(encoding='utf-8', errors='replace')

BASE = r"j:\Cosas Facu\Tp 5 Simulacion"
xlsx_path = BASE + r"\Trabajo Practico 5 - Grupo 10 - 4k3.xlsx"

wb = openpyxl.load_workbook(xlsx_path, data_only=True)
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n=== Hoja: '{sheet_name}' ({ws.max_row} filas x {ws.max_column} cols) ===")

    # Show merged cells info
    if ws.merged_cells.ranges:
        print(f"  Celdas fusionadas: {[str(m) for m in ws.merged_cells.ranges]}")

    for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
        if row_idx > 50:
            print(f"  ... ({ws.max_row - 50} filas mas)")
            break
        vals = [str(v) if v is not None else "" for v in row]
        # Only print if not all empty
        if any(v for v in vals):
            print(f"  R{row_idx:>3}: {' | '.join(vals)}")
wb.close()
